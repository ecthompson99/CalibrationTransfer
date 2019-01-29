#Importing modules

import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from matplotlib import pyplot as plt
from scipy.interpolate import interp1d
import numpy as np
from numpy import square as sq
from scipy.interpolate import InterpolatedUnivariateSpline as spl
from peakutils import indexes as peak


root = tk.Tk()
file_path = filedialog.askopenfilenames(filetype = [("Excel File","*.xlsx")])

first = load_workbook(str(file_path[0]))
second = load_workbook(str(file_path[1]))

sheet1 = first['ABS']
sheet2 = second['ABS']

# Fills up a list with an entire row
def fill_list(sheet, r, start = 1, length=sheet1.max_column):
    """ Fills up a list with an entire row from the excel sheet.
    
        Take two mandatory arguments, the sheetname and the row you want to scrape.
        
    """
    vals = []
    for i in range(start, length+1): 
        vals.append(sheet.cell(r, i).value)
    return vals

# Saving a 1D array of the x-values of each instrument
first_x = fill_list(sheet1, 1, 7)
first_x2 = fill_list(sheet2, 1, 7)

row = int(input('which row do you want to scrape?: '))

# saving a 1D array of the first set of y-values of each instrument (should correspond to the same sample)
first_row = fill_list(sheet1, row, 7)
first_row2 = fill_list(sheet2, row, 7)

interpolation_min = int(input('what is the lower limit of interpolation?: '))
interpolation_max = int(input('what is the upper limit of interpolation?: '))

x = np.linspace(interpolation_min,interpolation_max,251)

#interpolation: we can know all y-values between 450 and 700 nm. 
y1 = spl(first_x, first_row) # function
y2 = spl(first_x2, first_row2) # function

# interpolated y-values
yy1 = y1(x)
yy2 = y2(x)