#Importing modules

import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from matplotlib import pyplot as plt
from scipy.interpolate import interp1d
import numpy as np
from numpy import square as sq
from scipy.interpolate import InterpolatedUnivariateSpline as spl
from peakutils import indexes as peak


root = tk.Tk()
file_path = filedialog.askopenfilenames(filetype = [("Excel File","*.xlsx")])

first = load_workbook(str(file_path[0]))
second = load_workbook(str(file_path[1]))

sheet1 = first['ABS']
sheet2 = second['ABS']

# Fills up a list with an entire row
def fill_list(sheet, r, start = 1, length=sheet1.max_column):
    """ Fills up a list with an entire row from the excel sheet.
    
        Take two mandatory arguments, the sheetname and the row you want to scrape.
        
    """
    vals = []
    for i in range(start, length+1): 
        vals.append(sheet.cell(r, i).value)
    return vals

# Saving a 1D array of the x-values of each instrument
first_x = fill_list(sheet1, 1, 7)
first_x2 = fill_list(sheet2, 1, 7)

row = int(input('which row do you want to scrape?: '))

# saving a 1D array of the first set of y-values of each instrument (should correspond to the same sample)
first_row = fill_list(sheet1, row, 7)
first_row2 = fill_list(sheet2, row, 7)

interpolation_min = int(input('what is the lower limit of interpolation?: '))
interpolation_max = int(input('what is the upper limit of interpolation?: '))

x = np.linspace(interpolation_min,interpolation_max,251)

#interpolation: we can know all y-values between 450 and 700 nm. 
y1 = spl(first_x, first_row) # function
y2 = spl(first_x2, first_row2) # function

# interpolated y-values
yy1 = y1(x)
yy2 = y2(x)

################# MATHEMATICAL FUNCTIONS #######################


## SHIFTING FUNCTION ##
def shift():
    l = [] # array of the sum of squares
    
    shift_factor = np.linspace(-5, 5, 12)
    
    # Let's choose y1 as our master instrument
    # we need to shift y2 so that it has the smallest sum of squares with y1
    for i in shift_factor:
        nx = x + i
        ny2 = spl(nx, yy2) # same shape
        nyy2 = ny2(x) # new y-values
        l.append(sum(abs(sq(yy1)-sq(nyy2))))
    minum = [abs(num) for num in l]
    minindex = minum.index(min(minum))
    sf = list(shift_factor)[minindex]
    nx = x + sf
    ny2 = spl(nx, yy2)
    nyy2 = ny2(x)
    return yy1, nyy2, sf

y1, y2, sf = shift()


## TESTING THE SHIFTING FUNCTION BY PLOTTING ##

plt.figure(figsize=[20,20])
plt.subplot(2,1,1)
plt.plot(x, y1, x, y2)

plt.subplot(2,1,2)
plt.xlim(500, 650)
plt.ylim(0,1.2)
plt.plot(first_x, first_row, first_x2, first_row2)

## BANDWIDTH FUNCTION ##

peaks = peak(y2)

def bandwidth(peak_index, region): 
    lo = peak_index[0]-region
    hi = peak_index[-1]+region
    bandwidth = np.linspace(-10,10,201)
    p = []
    for k in bandwidth:
        l = []
        ny1 = []
        
        for i in range(lo, hi): 
            l.append(-y2[i-1]*k + (1+2*k)*y2[i] - y2[i+1]*k)
            ny1.append(y1[i])
        p.append(sum(abs(sq(l)-sq(ny1))))
        
    minum = [abs(num) for num in p]
    bw = list(bandwidth)[minum.index(min(minum))]
    ny2 = []
    for i in range(1, len(y2)-1): 
        ny2.append(-y2[i-1]*bw + (1+2*bw)*y2[i] - y2[i+1]*bw)
    return ny2, bw  

## TESTING THE BANDWIDTH FUNCTION BY PLOTTING ## 
y3, bw = bandwidth(peaks[1], 15)
plt.figure(figsize=[20,20])
x3 = np.linspace(501,649, 249)
plt.plot(x3, y3, 'g-', x, y1, 'b--')