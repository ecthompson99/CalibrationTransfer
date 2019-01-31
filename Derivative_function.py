#Importing modules

from tkinter import *
from tkinter import filedialog
from openpyxl import load_workbook
from matplotlib import pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from scipy.interpolate import interp1d, InterpolatedUnivariateSpline as spl
import numpy as np
from numpy import square as sq
from peakutils import indexes as peak
from scipy.signal import savgol_filter as sgf

root = Tk()
file_path = filedialog.askopenfilenames(filetype = [("Excel File","*.xlsx")])

first = load_workbook(str(file_path[0]))
second = load_workbook(str(file_path[1]))

sheet1 = first['ABS']
sheet2 = second['ABS']

# Fills up a list with an entire row
def fill_list(sheet, r, start = 1, length=sheet1.max_column):
    vals = []
    for i in range(start, length+1): 
        vals.append(sheet.cell(r, i).value)
    return vals

# Saving a 1D array of the x-values of each instrument

# def find_rcal():
#     indexes = []
#     for i in range(1, sheet1.max_column):
#         if 'RCAL' in str(sheet1.cell(i, 1).value):
#             indexes.append(i)
#     return indexes

# rcal_i = find_rcal() # using the rcal samples
rcal_i = [24, 25, 26, 27, 28, 29] # uses the samples corresponding to the rows in the excel file

first_x = fill_list(sheet1, 1, 7) 
first_x2 = fill_list(sheet2, 1, 7) 

first_row = [] # stores 3 sets of y-values for the calibration samples of the main instrument
first_row2 = [] # stores 3 sets of y-values for the calibration samples of the secondary instrument

y1 = []
y2 = []
yy1 = []
yy2 = []

x = [first_x[i] for i in range(50, 85)]

for i in rcal_i:
    first_row.append(fill_list(sheet1, i, 7))
    first_row2.append(fill_list(sheet2, i, 7))

for i in range(len(rcal_i)):
    y1.append(spl(first_x, first_row[i]))
    y2.append(spl(first_x2, first_row2[i]))
    yy1.append(y1[i](x))
    yy2.append(y2[i](x))

yyy1 = []
yyy2 = []

for i in range(len(rcal_i)):
    yyy1.append((sgf(yy1[i], 7, 3, deriv=1)))
    yyy2.append((sgf(yy2[i], 7, 3, deriv=1)))

def shift(yy1, yy2):
    l = [] # array of the sum of squares
    shift_factor = np.linspace(-5, 5, 1001)

    for i in shift_factor:
        nx = x + i
        nyy2 = list(np.ones(len(yy2)))
        squaresum = []

        for j in range(len(yy2)): 
            ny2 = spl(nx, yy2[j]) # same shape
            nyy2[j] = ny2(x) # new y-values

            squaresum.append(sum(abs(sq(yy1[j])-sq(nyy2[j]))))

        l.append(sum(squaresum))

    minindex = l.index(min(l))
    sf = list(shift_factor)[minindex]
    nx = x + sf
    nyy2 = []
    for i in range(len(yy2)): 
        ny2 = spl(nx, yy2[i])
        nyy2.append(ny2(x))

    return nyy2, sf

nyy2, sf = shift(yyy1, yyy2)

peaks = [peak(nyy2[i]) for i in range(len(nyy2))]

x3 = [x[i] for i in range(1,len(x)-1)]

def bandwidth(peak_index, region, nyy2, yy1):

    bandwidth = np.linspace(-5,5,1001)
    p = []

    for k in bandwidth:
        l2 = []
        ny2 = []

        for i in range(len(peak_index)): 
            lo = peak_index[i][0]-region
            hi = peak_index[i][-1]+region
            l = []
            ny1 = []

            for j in range(lo, hi): 
                l.append(-nyy2[i][j-1]*k + (1+2*k)*nyy2[i][j] - nyy2[i][j+1]*k)
                ny1.append(yy1[i][j])

            l2.append(l)
            ny2.append(ny1)            

        sumsquare = [sum(abs(sq(l2[i])-sq(ny2[i]))) for i in range(len(peak_index))]
        p.append(sum(sumsquare))

    bw = list(bandwidth)[p.index(min(p))]
    bwy2 = []
    for i in range(len(peak_index)):
        temp = []
        for j in range(1, len(x)-1):
            temp.append(-nyy2[i][j-1]*bw + (1+2*bw)*nyy2[i][j] - nyy2[i][j+1]*bw)
        bwy2.append(temp)

    return bwy2, bw, min(p)

bwy2, bw, err = bandwidth(peaks, 1, nyy2, yyy1)

from scipy.integrate import cumtrapz
x4 = [x[i] for i in range(1, len(x))]

new_y2 = []
old_y1 = []
sumsquare = []
shrink = 2;
for i in range(len(bwy2)): 
    new_y2.append(cumtrapz(nyy2[i]/shrink, x))
    old_y1.append(cumtrapz(yyy1[i]/shrink, x))
    sumsquare.append(sum(abs(sq(new_y2[i])-sq(old_y1[i]))))

plt.cla()

fig = plt.figure()

# # DERIVATIVE PLOTS
# fig.text(0.5, .05, "The difference of squares error is: " + str(err), ha='center')
# plt.plot(x3, bwy2[0], 'g-', x, yyy1[0], 'b-', x, yyy2[0], 'r-')
# plt.plot(x3, bwy2[1], 'g-', x, yyy1[1], 'b-', x, yyy2[1], 'r-')
# plt.plot(x3, bwy2[2], 'g-', x, yyy1[2], 'b-', x, yyy2[2], 'r-')
# plt.plot(x3, bwy2[3], 'g-', x, yyy1[3], 'b-', x, yyy2[3], 'r-')
# plt.plot(x3, bwy2[4], 'g-', x, yyy1[4], 'b-', x, yyy2[4], 'r-')
# plt.plot(x3, bwy2[5], 'g-', x, yyy1[5], 'b-', x, yyy2[5], 'r-')
# plt.title('Derivatized plot of calibration with derivative function')


# INTEGRAL PLOTS
fig.text(0.5, .05, "The difference of squares error is: " + str(sum(sumsquare)), ha='center')
plt.plot(x4, new_y2[0], 'g-', x4, old_y1[0], 'b-')
plt.plot(x4, new_y2[1], 'g-', x4, old_y1[1], 'b-')
plt.plot(x4, new_y2[2], 'g-', x4, old_y1[2], 'b-')
plt.plot(x4, new_y2[3], 'g-', x4, old_y1[3], 'b-')
plt.plot(x4, new_y2[4], 'g-', x4, old_y1[4], 'b-')
plt.plot(x4, new_y2[5], 'g-', x4, old_y1[5], 'b-')
plt.title('Integrated plot of calibration with derivative function')

print(sf, bw)
plt.show()