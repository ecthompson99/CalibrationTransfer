#Importing modules

import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from matplotlib import pyplot as plt
from scipy.interpolate import interp1d
import numpy as np
from numpy import square as sq

root = tk.Tk()
file_path = filedialog.askopenfilenames(filetype = [("Excel File","*.xlsx")])

### loading workbooks and sheets

first = load_workbook(str(file_path[0]))
second = load_workbook(str(file_path[1]))

sheet1 = first['ABS']
sheet2 = second['ABS']

### scraping data
# Fills up a list with an entire row
def fill_list(sheet, r, start = 1, length=sheet1.max_column):
    """ Fills up a list with an entire row from the excel sheet.
    
        Take two mandatory arguments, the sheetname and the row you want to scrape.
        
    """
    vals = []
    for i in range(start, length+1): 
        vals.append(sheet.cell(r, i).value)
    return vals

first_x = fill_list(sheet1, 1, 7)
first_x2 = fill_list(sheet2, 1, 7)

first_row = fill_list(sheet1, 2, 7)
first_row2 = fill_list(sheet2, 2, 7)

def interp(values, x, lo = 450, hi = 851):
    sheet1y = []
    for i in range(len(values)):
        f = interp1d(x,values[i])
        sheet1y.append(f(range(lo, hi)))
    return sheet1y

f1 = interp1d(first_x, first_row)
f2 = interp1d(first_x2, first_row2)

interp1 = [f1(i) for i in np.linspace(450, 850, 401)]
interp2 = [f2(i) for i in np.linspace(450, 850, 401)]

x = np.linspace(450, 850, 401)

def shift(lo = -5, hi = 5, numvals = 100, region = [500, 600]):
    l = []

    for i in np.linspace(lo,hi,numvals):
        nx = x+i
        testx = np.linspace(region[0],region[1],1+region[1]-region[0])
        t = interp1d(nx, interp1)
        y = [t(i) for i in testx]
        testy = [f2(i) for i in testx]
        l.append(sum(sq(y))-sum(sq(testy)))

    minum = min([abs(num) for num in l])
    nl = [abs(num) for num in l]
    sf = np.linspace(lo,hi,numvals)[nl.index(minum)]
    return sf

sf = shift()

def bandwidth():
    nx = x + sf
    p = []
    testx = np.linspace(500, 600, 101)
    testx2 = np.linspace(498, 601, 104)
    testy = [f2(i) for i in testx]
    blah = np.linspace(-2,2,100)
    for k in blah:
        l = [] 
        t = interp1d(nx, interp1)
        y = [t(i) for i in testx2]

        for i in range(2, len(testx2)-1): 
            l.append(-y[i-1]*k + (1+2*k)*y[i] - y[i+1]*k)

        p.append(sum(sq(l)-sq(testy)))

    minum = min([abs(num) for num in p])
    pp = [abs(num) for num in p]
    bw = blah[pp.index(minum)]
    return bw

# plt.figure(figsize=[30,30])
# plt.xlim([500,600])
# plt.ylim([1, 1.4])
# plt.plot(x, interp2, '--', x, interp1, '-*', x+sf, interp1, '-')