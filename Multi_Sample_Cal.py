#Importing modules

from tkinter import *
from tkinter import filedialog
from openpyxl import load_workbook
from matplotlib import pyplot as plt
from scipy.interpolate import interp1d, InterpolatedUnivariateSpline as spl
import numpy as np
from numpy import square as sq
from peakutils import indexes as peak

# Tkinter allows you to open a file
root = Tk()
file_path = filedialog.askopenfilenames(filetype = [("Excel File","*.xlsx")])

first = load_workbook(str(file_path[0]))     # The Master instrument is assumed to be the first file
second = load_workbook(str(file_path[1]))    # The second instrument

# Selects the absorbance sheet. Make sure the name hasn't changed.
sheet1 = first['ABS']
sheet2 = second['ABS']

# Function that allows you to scrape a row. The user must specify which column to start at.
def fill_list(sheet, r, start = 1, length=sheet1.max_column):
    vals = []
    for i in range(start, length+1): 
        vals.append(sheet.cell(r, i).value)
    return vals

# picking the interval of rows of calibration samples
rcal_i = [i for i in range(77,83)] # for shift; first value included, second value excluded

# scrapes the x-values for the master and slave instruments
first_x = fill_list(sheet1, 1, 7) 
first_x2 = fill_list(sheet2, 1, 7) 

# An empty array to be filled with the spectra of the selected samples
first_row = [] # from the master instrument
first_row2 = [] # from the slave instrument

# Empty arrays to be filled with the splines (y1, y2), and the interpolated y-values (yy1, yy2)
y1 = [] # master instrument splines
y2 = []
yy1 = [] # master instrument interpolated y-values
yy2 = []

# for broader peaks a range(40, 105) covers 400 nm to 700 nm
# for narrower peaks a range of (50, 90) covers 500 nm to 640 nm
x = [first_x[i] for i in range(40, 105)] # Here, the range of y-values are chosen

for i in rcal_i:
    first_row.append(fill_list(sheet1, i, 7))
    first_row2.append(fill_list(sheet2, i, 7))

for i in range(len(rcal_i)):
    y1.append(spl(first_x, first_row[i]))
    y2.append(spl(first_x2, first_row2[i]))
    yy1.append(y1[i](x))
    yy2.append(y2[i](x))
    
# -------------------- FUNCTIONS -----------------------------------

### Mean subtraction

mean_yy1 = [np.mean(yy1[i]) for i in range(len(yy1))]
mean_yy2 = [np.mean(yy2[i]) for i in range(len(yy2))]
sm_yy1 = []
sm_yy2 = []
for i in range(len(yy1)): 
    sm_yy1.append([yy1[i][j] - mean_yy1[i] for j in range(len(yy1[i]))])
    sm_yy2.append([yy2[i][j] - mean_yy2[i] for j in range(len(yy1[i]))])

    
### shifting function

def shift(yy1, yy2):
    l = [] # array of the sum of squares
    shift_factor = np.linspace(-5, 5, 1001)
    
    for i in shift_factor:
        nx = x + i
        nyy2 = list(np.ones(len(yy2)))
        squaresum = []
        
        for j in range(len(yy2)): 
            ny2 = spl(nx, yy2[j]) # same shape
            nyy2[j] = ny2(x) # new y-values
            
            squaresum.append(sum(abs(yy1[j]-nyy2[j])))
          
        l.append(sum(squaresum))

    minindex = l.index(min(l))
    sf = list(shift_factor)[minindex]
    nx = x + sf
    nyy2 = []
    for i in range(len(yy2)): 
        ny2 = spl(nx, yy2[i])
        nyy2.append(ny2(x))
   
    return nyy2, sf

nyy2, sf = shift(sm_yy1, sm_yy2)


### Bandwidth Function

peaks = [peak(nyy2[i]) for i in range(len(nyy2))]
x3 = [x[i] for i in range(1,len(x)-1)]

def bandwidth(peak_index, region, nyy2, yy1):
    bandwidth = np.linspace(-5,5,1001)
    p = []
    
    for k in bandwidth:
        l2 = []
        ny2 = []
        
        for i in range(len(peak_index)): 
            lo = peak_index[i][0]-region
            hi = peak_index[i][-1]+region
            l = []
            ny1 = []

            for j in range(lo, hi): 
                l.append(-nyy2[i][j-1]*k + (1+2*k)*nyy2[i][j] - nyy2[i][j+1]*k)
                ny1.append(yy1[i][j])
            
            l2.append(l)
            ny2.append(ny1)            
        
        sub_arrays = [[abs(a - b) for (a, b) in zip(l2[i], ny2[i])] for i in range(len(peak_index))]
        sumsquare = [sum(num) for num in sub_arrays]
        p.append(sum(sumsquare))

    bw = list(bandwidth)[p.index(min(p))]
    bwy2 = []
    for i in range(len(peak_index)):
        temp = []
        for j in range(1, len(x)-1):
            temp.append(-nyy2[i][j-1]*bw + (1+2*bw)*nyy2[i][j] - nyy2[i][j+1]*bw)
        bwy2.append(temp)
        
    return bwy2, bw, min(p)

bwy2, bw, err = bandwidth(peaks, 10, nyy2, sm_yy1)

    
# ------------------------- PLOTS --------------------------------------

%matplotlib qt

fig = plt.figure()

# You can plot up to 6 curves at a time (you can choose to eliminate certain curves by commenting out the corresponding line).
plt.plot(x3, bwy2[0], 'g-', x, sm_yy1[0], 'b-', x, sm_yy2[0], 'r-')
# plt.plot(x3, bwy2[1], 'g-', x, sm_yy1[1], 'b-', x, sm_yy2[1], 'r-')
plt.plot(x3, bwy2[2], 'g-', x, sm_yy1[2], 'b-', x, sm_yy2[2], 'r-')
# plt.plot(x3, bwy2[3], 'g-', x, sm_yy1[3], 'b-', x, sm_yy2[3], 'r-')
plt.plot(x3, bwy2[4], 'g-', x, sm_yy1[4], 'b-', x, sm_yy2[4], 'r-')
# plt.plot(x3, bwy2[5], 'g-', x, sm_yy1[5], 'b-', x, sm_yy2[5], 'r-')

plt.title('Transformed Slave Spectrum')
plt.show()

print('Shift: {} \nBandwidth {}'.format(round(sf, 3), round(bw, 3)))
